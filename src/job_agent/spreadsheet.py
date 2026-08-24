from __future__ import annotations

from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import JobLead
from .resume import ResumeStatus


HEADERS = [
    "Rank",
    "Job Title",
    "Company",
    "Location",
    "Experience",
    "Work Mode",
    "Match Score",
    "Apply Link",
    "Source Platform",
    "Posted Date",
    "Short Job Summary",
    "Auto Apply Eligible",
    "Auto Apply Reason",
    "Application Status",
    "Tailored Pitch",
    "Cover Note",
    "Suggested Answers",
]


def export_jobs_xlsx(jobs: list[JobLead], output_dir: Path, resume_status: ResumeStatus | None = None) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(ZoneInfo("Asia/Kolkata")).strftime("%Y%m%d_%H%M%S_IST")
    output_path = output_dir / f"job_search_top_20_{timestamp}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Top 20 Jobs"

    sheet.append(HEADERS)
    for rank, job in enumerate(jobs[:20], start=1):
        sheet.append(
            [
                rank,
                job.title,
                job.company,
                job.location,
                job.experience,
                job.work_mode,
                round(job.match_score, 1),
                job.apply_link,
                job.source_platform,
                job.posted_date,
                job.short_summary,
                "Yes" if job.auto_apply_eligible else "No",
                job.auto_apply_reason,
                job.application_status,
                job.tailored_pitch,
                job.cover_note,
                job.suggested_answers,
            ]
        )

    style_sheet(sheet, len(jobs[:20]) + 1)
    add_metadata_sheet(workbook, jobs, resume_status)
    workbook.save(output_path)
    return output_path


def style_sheet(sheet, last_row: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 8,
        "B": 34,
        "C": 24,
        "D": 24,
        "E": 16,
        "F": 16,
        "G": 13,
        "H": 52,
        "I": 18,
        "J": 14,
        "K": 60,
        "L": 18,
        "M": 42,
        "N": 22,
        "O": 48,
        "P": 60,
        "Q": 60,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2, max_row=max(last_row, 2), max_col=len(HEADERS)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        score = row[6].value
        if isinstance(score, (int, float)):
            if score >= 80:
                row[6].fill = PatternFill("solid", fgColor="C6EFCE")
            elif score >= 65:
                row[6].fill = PatternFill("solid", fgColor="FFEB9C")
            else:
                row[6].fill = PatternFill("solid", fgColor="F8CBAD")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(last_row, 1)}"

    if last_row >= 2:
        table = Table(displayName="TopJobs", ref=f"A1:{get_column_letter(len(HEADERS))}{last_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)


def add_metadata_sheet(workbook: Workbook, jobs: list[JobLead], resume_status: ResumeStatus | None = None) -> None:
    sheet = workbook.create_sheet("Run Metadata")
    metadata = [
        ("Generated At IST", datetime.now(ZoneInfo("Asia/Kolkata")).isoformat(timespec="seconds")),
        ("Result Count", min(len(jobs), 20)),
        ("Freshness Rule", "Only jobs with posted dates within the last 3 days are included."),
        ("Experience Rule", "Entry-level and roles requiring no more than 2 years are included."),
        ("Automation Boundary", "Public indexed pages only. No authenticated pages, no auto-apply, no CAPTCHA bypass."),
        ("Auto Apply Policy", "Job-board applications are review-only. Public company/ATS pages are only eligible when explicitly allowlisted."),
    ]
    if resume_status:
        metadata.extend(
            [
                ("Resume Status", resume_status.status),
                ("Resume Source", resume_status.source),
                ("Resume Reference", resume_status.reference),
            ]
        )
    for row in metadata:
        sheet.append(row)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 90
    for cell in sheet[1]:
        cell.font = Font(bold=True)
